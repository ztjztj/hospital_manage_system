from django.shortcuts import render,redirect,reverse
from registration.models import *
from django.core.paginator import Paginator
from django.http import HttpResponse
from doctor.forms import *
import openpyxl


# 主页

def index(request):
    return render(request,'doctor/index.html')



# 信息查询
def find_0(request):
    pname = request.POST.get("pname")
    request.session["id"] = pname
    return redirect("query")


def query(request):
    id = request.session.get("id")
    if id:
        if PatientOne.objects.filter(patient_status=id).exists():
            messages_1 = PatientOne.objects.values("patient_status").distinct()
            messages = PatientOne.objects.filter(patient_status=id,patient_over_look_doctor=0)
            print(messages)
            limit =3
            paginator = Paginator(messages, limit)  # 按每页10条分页
            page = request.GET.get('page', '1')  # 默认跳转到第一页
            result = paginator.page(page)
            return render(request, 'doctor/patient_list.html', {"messages_1": messages_1,"messages": result})

        elif PatientOne.objects.filter(patient_name=id).exists():
            messages_1 = PatientOne.objects.values("patient_status").distinct()
            messages = PatientOne.objects.filter(patient_name=id,patient_over_look_doctor=0)
            limit = 3
            paginator = Paginator(messages, limit)  # 按每页10条分页
            page = request.GET.get('page', '1')  # 默认跳转到第一页
            result = paginator.page(page)
            return render(request, 'doctor/patient_list.html', {"messages_1": messages_1,"messages": result})
        elif PatientOne.objects.filter(patient_card=id).exists():
            messages_1 = PatientOne.objects.values("patient_status").distinct()
            messages = PatientOne.objects.filter(patient_card=id,patient_over_look_doctor=0)
            limit = 3
            paginator = Paginator(messages, limit)  # 按每页10条分页
            page = request.GET.get('page', '1')  # 默认跳转到第一页
            result = paginator.page(page)
            return render(request, 'doctor/patient_list.html', {"messages_1": messages_1,"messages": result})
        else:
            messages_1 = PatientOne.objects.values("patient_status").distinct()
            messages = PatientOne.objects.filter(patient_over_look_doctor=0)
            limit = 3
            paginator = Paginator(messages, limit)  # 按每页10条分页
            page = request.GET.get('page', '1')  # 默认跳转到第一页
            result = paginator.page(page)
            return render(request, 'doctor/patient_list.html', {"messages_1": messages_1,"messages": result})
    else:
        messages_1 = PatientOne.objects.values("patient_status").distinct()
        messages = PatientOne.objects.filter(patient_over_look_doctor=0)
        limit = 3
        paginator = Paginator(messages, limit)  # 按每页10条分页
        page = request.GET.get('page', '1')  # 默认跳转到第一页
        result = paginator.page(page)
        return render(request, 'doctor/patient_list.html', {"messages_1": messages_1,"messages": result})



# 指定病人病症的信息的查询并分页
def patient_list_1(request,id):

    if id:
        messages_1 = PatientOne.objects.values("patient_status").distinct()
        messages = PatientOne.objects.filter(patient_status=id,patient_over_look_doctor=0)
        print(messages)
        limit = 3
        paginator = Paginator(messages, limit)  # 按每页10条分页
        page = request.GET.get('page', '1')  # 默认跳转到第一页
        result = paginator.page(page)
        return render(request, 'doctor/patient_list.html', {"messages_1": messages_1,"messages": result})
    else:
        messages_1 = PatientOne.objects.values("patient_status").distinct()
        messages = PatientOne.objects.filter(patient_over_look_doctor=0)
        limit = 3
        paginator = Paginator(messages, limit)  # 按每页10条分页
        page = request.GET.get('page', '1')  # 默认跳转到第一页
        result = paginator.page(page)
        return render(request, 'doctor/patient_list.html', {"messages_1": messages_1, "messages": result})





# 所有病人信息的查询并分页

def patient_list(request):
    messages_1 = PatientOne.objects.values("patient_status").distinct()
    messages = PatientOne.objects.filter(patient_over_look_doctor=0)
    limit = 3
    paginator = Paginator(messages, limit)  # 按每页10条分页
    page = request.GET.get('page', '1')  # 默认跳转到第一页
    result = paginator.page(page)
    return render(request, 'doctor/patient_list.html', {"messages_1": messages_1,"messages": result})




# 病症详情

def detail(request,id):
    详情 = PatientOne.objects.get(id=id)
    return render(request,"doctor/detail.html",{"详情":详情})




# 就诊

def dispensing(request,id):
    us_er = PatientOne.objects.get(id=id)
    a= us_er.patient_is_live
    print(type(a))
    yao_pin = Medicine.objects.all()  # 后期需要修改
    medicine_one = Medicine_one.objects.filter(fk_patient_medicine_one=id)
    paginator = Paginator(yao_pin, 5)  # 按每页10条分页
    page = request.GET.get('page', '1')  # 默认跳转到第一页
    result = paginator.page(page)
    paginator_1 = Paginator(medicine_one, 5)  # 按每页10条分页
    page_1 = request.GET.get('page_1', '1')  # 默认跳转到第一页
    result_1 = paginator_1.page(page_1)
    return render(request,"doctor/dispensing.html",{"us_er":us_er,"yao_pin":result,"yao_fang":result_1})










# 开药方
def Medicine_one_add(request,id):
    if request.method == "GET":
        us_er = PatientOne.objects.get(id=id)
        yao_pin = Medicine.objects.all()  # 后期需要修改
        medicine_one = Medicine_one.objects.filter(fk_patient_medicine_one=id)
        limit = 5
        paginator = Paginator(yao_pin, limit)  # 按每页10条分页
        page = request.GET.get('page', '1')  # 默认跳转到第一页
        result = paginator.page(page)
        return render(request, "doctor/dispensing.html", {"us_er": us_er, "yao_pin": result})
    else:
        # print("*" * 100000)
        medicine_name = request.POST.get("ypmz")
        print("*" * 100000)
        if Medicine.objects.filter(medicine_name=medicine_name,).exists() and not Medicine_one.objects.filter(medicine_one_name=medicine_name,medicine_one_yes_no=1):
            # print("*" * 100000)
            patient_one1 = Medicine.objects.get(medicine_name=medicine_name)
            patient_one = patient_one1.medicine_outer_price
            # print(id)
            print(medicine_name)
            # print(patient_one)
            medicine_number = request.POST.get("ypsl")
            if medicine_number and 50>int(medicine_number)>0:
                # print(medicine_number)
                medicine_one = Medicine_one.objects.create(medicine_one_name=medicine_name,
                                                           medicine_one_number=medicine_number,
                                                           medicine_one_outer_price=patient_one,
                                                           fk_patient_medicine_one_id=id, )
                # print("*" * 100000)
                return redirect(reverse(dispensing, kwargs={"id": id}))
            else:
                return redirect(reverse(dispensing, kwargs={"id": id}))
        else:
            return redirect(reverse(dispensing,kwargs={"id":id}))


# 撤回药方中的药品

def delete_yaofan(request,id):
    user = Medicine_one.objects.get(medicine_one_name=id)
    if id:
        delete_1 = Medicine_one.objects.get(medicine_one_name=id)
        delete_1.delete()
        id = user.fk_patient_medicine_one_id
        return redirect(reverse(dispensing, kwargs={"id": id}))
    else:
        return redirect(reverse(dispensing, kwargs={"id": id}))

# # 住院

def hospital(request,id):
    print(id)
    if request.method == "GET":
        us_er = PatientOne.objects.get(id=id)
        yao_pin = Medicine.objects.all()  # 后期需要修改
        medicine_one = Medicine_one.objects.filter(fk_patient_medicine_one=id)
        limit = 5
        paginator = Paginator(yao_pin, limit)  # 按每页10条分页
        page = request.GET.get('page', '1')  # 默认跳转到第一页
        result = paginator.page(page)
        return render(request, "doctor/dispensing.html", {"us_er": us_er, "yao_pin": result})
    else:
        user = PatientOne.objects.get(id=id)
        day = request.POST.get("day")
        print(day)
        if 0<=int(day)<10000 and day:    # 缺少条件，先住再续
            user.patient_is_live = 1
            user.patient_live_day = day
            print(user.patient_is_live)
            user.save()
            return redirect(reverse(dispensing, kwargs={"id": id}))
        else:
            return redirect(reverse(dispensing, kwargs={"id": id}))




# # 药方输出到excle中

def save_excle(request,id):
    medicine_one = Medicine_one.objects.filter(fk_patient_medicine_one=id)
    for i in medicine_one:
        print("yaoming",i.medicine_one_name)
        print("yaojia",i.medicine_one_outer_price)
        print("yaoshu",i.medicine_one_number)
        print("*"*100000)
    print("*" * 100000)
    patientone = PatientOne.objects.get(id=id)
    print("bingren",patientone.patient_name)
    print("bingzheng",patientone.patient_status)
    print("zhuzhidaifu",patientone.patient_main_doctor)
    print("day",patientone.patient_live_day)


    print("*" * 100000)


    one = patientone.patient_name
    a = openpyxl.Workbook()  # 打开文件
    b = a.active  # 激活表格
    b.title = "%s" % one  # 添加工作簿名称
    e = [[ "病人", "病症", "主治大夫","住院天数"],
         [patientone.patient_name,patientone.patient_status,patientone.patient_main_doctor,patientone.patient_live_day],
         ]
    for i in range(len(e)):
        for j in range(len(e[i])):
            b.cell(i + 1, j + 1, e[i][j])  # 写入单元格
    e1 = [["药名", "价钱", "数量", ]]
    for i in medicine_one:
        f = [i.medicine_one_name, i.medicine_one_outer_price, i.medicine_one_number]  # 添加内容
        e1.append(f)
        for i in range(len(e1)):
            for j in range(len(e1[i])):
                b.cell(i + 1, j + 5, e1[i][j])  # 写入单元格
    a.save(filename='d:\\one_1.xlsx')  # 保存文件
    a.close()
    return redirect(reverse(dispensing, kwargs={"id": id}))



# 提交并返回主页

def submit(request,id):
    user = PatientOne.objects.get(id=id)
    medicine = Medicine_one.objects.filter(fk_patient_medicine_one=id)
    if medicine:
        medicine_one_yes_no = Medicine_one.objects.filter(fk_patient_medicine_one=id)
        medicine_one_yes_no.update(medicine_one_yes_no=1)
        print(medicine_one_yes_no)
        user.patient_over_look_doctor = 1
        user.save()
        return redirect("patient_list")
    else:
        return redirect(reverse(dispensing, kwargs={"id": id}))


















# 密码
def doctor_password(request):
    return render(request,'doctor/doctor_password.html')


